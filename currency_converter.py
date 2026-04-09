"""
Currency Converter
Converts all currency pairs to nominal USD amounts
"""

import requests
from typing import Dict, List
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os


class CurrencyConverter:
    """A class to handle currency conversion to USD"""
    
    BASE_URL = "https://api.exchangerate-api.com/v4/latest/"
    
    def __init__(self):
        """Initialize the currency converter"""
        self.rates = {}
        self.last_updated = None
    
    def fetch_rates(self, currency: str = "USD") -> Dict[str, float]:
        """
        Fetch exchange rates for a given currency
        
        Args:
            currency: Currency code (default: USD)
        
        Returns:
            Dictionary of currency pairs and their rates
        """
        try:
            response = requests.get(f"{self.BASE_URL}{currency}")
            response.raise_for_status()
            data = response.json()
            self.rates = data.get("rates", {})
            self.last_updated = datetime.now()
            return self.rates
        except requests.RequestException as e:
            print(f"Error fetching rates: {e}")
            return {}
    
    def convert_to_usd(self, amount: float, from_currency: str) -> float:
        """
        Convert an amount from a given currency to USD
        
        Args:
            amount: Amount to convert
            from_currency: Currency code to convert from
        
        Returns:
            Amount in USD
        """
        if not self.rates:
            self.fetch_rates(from_currency)
        
        if from_currency == "USD":
            return amount
        
        # If we fetched rates for a different currency, get USD rates instead
        if from_currency not in self.rates:
            self.fetch_rates("USD")
        
        # Use inverse rate to convert to USD
        if from_currency in self.rates:
            rate = self.rates[from_currency]
            return amount / rate
        else:
            print(f"Currency {from_currency} not found")
            return 0.0
    
    def convert_multiple(self, pairs: List[tuple]) -> Dict[str, float]:
        """
        Convert multiple currency pairs to USD
        
        Args:
            pairs: List of tuples (amount, currency_code)
        
        Returns:
            Dictionary with converted USD amounts
        """
        self.fetch_rates("USD")
        results = {}
        
        for amount, currency in pairs:
            results[f"{amount} {currency}"] = self.convert_to_usd(amount, currency)
        
        return results
    
    def get_conversion_rate(self, from_currency: str, to_currency: str = "USD") -> float:
        """
        Get the conversion rate between two currencies
        
        Args:
            from_currency: Source currency code
            to_currency: Target currency code (default: USD)
        
        Returns:
            Conversion rate
        """
        if not self.rates:
            self.fetch_rates("USD")
        
        if to_currency == "USD":
            if from_currency in self.rates:
                return 1 / self.rates[from_currency]
            else:
                print(f"Currency {from_currency} not found")
                return 0.0
        
        return 0.0
    
    def process_excel_file(self, file_path: str, amount_col: str = "Amount", 
                          currency_col: str = "Currency", output_file: str = None) -> Dict:
        """
        Process an Excel file and convert currencies to USD
        
        Args:
            file_path: Path to the Excel file
            amount_col: Column name for amounts (default: "Amount")
            currency_col: Column name for currencies (default: "Currency")
            output_file: Path to save results (default: adds _USD to input filename)
        
        Returns:
            Dictionary with conversion results
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Validate columns exist
            if amount_col not in df.columns or currency_col not in df.columns:
                print(f"Error: Columns '{amount_col}' or '{currency_col}' not found in Excel file")
                print(f"Available columns: {list(df.columns)}")
                return {}
            
            # Fetch rates once
            self.fetch_rates("USD")
            
            # Convert each row
            df['USD_Amount'] = df.apply(
                lambda row: self.convert_to_usd(float(row[amount_col]), str(row[currency_col])),
                axis=1
            )
            df['USD_Amount'] = df['USD_Amount'].round(2)
            
            # Add conversion rate column
            df['Exchange_Rate'] = df[currency_col].apply(
                lambda x: round(self.get_conversion_rate(str(x)), 4)
            )
            
            # Save results
            if output_file is None:
                base, ext = os.path.splitext(file_path)
                output_file = f"{base}_USD.xlsx"
            
            self._save_excel_with_formatting(df, output_file)
            
            print(f"✓ Conversion complete!")
            print(f"✓ Results saved to: {output_file}")
            
            return {
                "file": output_file,
                "rows_processed": len(df),
                "data": df.to_dict('records')
            }
        
        except FileNotFoundError:
            print(f"Error: File '{file_path}' not found")
            return {}
        except Exception as e:
            print(f"Error processing Excel file: {e}")
            return {}
    
    def _save_excel_with_formatting(self, df: pd.DataFrame, file_path: str):
        """
        Save DataFrame to Excel with formatting
        
        Args:
            df: DataFrame to save
            file_path: Output file path
        """
        df.to_excel(file_path, index=False, sheet_name='Conversions')
        
        # Add formatting
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Currency column (yellow highlight)
        currency_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Amount column (light green)
        amount_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # USD column (light blue)
        usd_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        
        # Apply column widths and formats
        ws.column_dimensions['A'].width = 15
        for col_num, column_title in enumerate(df.columns, 1):
            ws.column_dimensions[chr(64 + col_num)].width = 18
            
            # Color code columns
            for row_num in range(2, len(df) + 2):
                cell = ws.cell(row=row_num, column=col_num)
                cell.number_format = '0.00' if 'Amount' in column_title or 'Rate' in column_title else '@'
        
        wb.save(file_path)


def main():
    """Main function to demonstrate currency conversion"""
    converter = CurrencyConverter()
    
    print("="*60)
    print("CURRENCY CONVERTER - Convert All Currencies to USD")
    print("="*60)
    print()
    
    # Check if sample Excel file exists
    sample_file = "sample_currencies.xlsx"
    if os.path.exists(sample_file):
        print(f"Found Excel file: {sample_file}")
        print("Processing currencies from Excel file...")
        print("-" * 60)
        result = converter.process_excel_file(sample_file)
        if result:
            print()
            print("Converted Data:")
            for item in result.get('data', [])[:5]:  # Show first 5 rows
                print(f"  {item}")
        print()
    else:
        print(f"Note: No Excel file found. Create '{sample_file}' with columns:")
        print("  - 'Amount': numeric amount to convert")
        print("  - 'Currency': currency code (EUR, GBP, JPY, etc.)")
        print()
    
    # Example usage
    print("Manual Conversion Examples:")
    print("-" * 60)
    
    # Fetch and display some rates
    converter.fetch_rates("USD")
    print(f"Last updated: {converter.last_updated}")
    print()
    
    # Convert single amounts
    test_pairs = [
        (100, "EUR"),
        (1000, "JPY"),
        (50, "GBP"),
        (200, "CAD"),
    ]
    
    print("Converting various amounts to USD:")
    for amount, currency in test_pairs:
        usd_amount = converter.convert_to_usd(amount, currency)
        print(f"  {amount} {currency}: ${usd_amount:.2f}")
    
    print()
    print("Conversion rates to USD:")
    print("-" * 60)
    
    currencies = ["EUR", "JPY", "GBP", "CAD"]
    for currency in currencies:
        rate = converter.get_conversion_rate(currency)
        if rate > 0:
            print(f"  1 {currency} = ${rate:.4f}")


if __name__ == "__main__":
    main()
