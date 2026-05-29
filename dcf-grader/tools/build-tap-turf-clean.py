"""Build the Tap & Turf completed DCF as a CLEAN workbook from scratch.

The supplied template ("...Corrupt.xlsx") carries FactSet/Bloomberg add-in
cruft and won't open reliably once round-tripped. Rather than patch a broken
file, we generate a fresh, self-contained workbook — guaranteed-valid OOXML —
with the same tab structure (Assumptions / Company fin forecasts / DCF).

All forecast cells are live formulas linking to the Assumptions driver block,
so changing one assumption recalculates the whole model.

Run from repo root:
    python3 dcf-grader/tools/build-tap-turf-clean.py \\
        --dst dcf-grader/sample-models/tap-turf-completed.xlsx
"""

import argparse
import importlib.util
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- styling helpers -------------------------------------------------------
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="4472C4")
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FONT = Font(color="0000CC")           # blue = input (banker convention)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0.0;(#,##0.0)"
PCT = "0.0%"
MULT = '0.00"x"'


def load_dcf():
    here = pathlib.Path(__file__).resolve().parent
    spec = importlib.util.spec_from_file_location("tap_turf_dcf", here / "tap-turf-dcf.py")
    m = importlib.util.module_from_spec(spec)
    sys.modules["tap_turf_dcf"] = m
    spec.loader.exec_module(m)
    return m


def style_title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = TITLE


def hdr(ws, cell, text):
    ws[cell] = text
    ws[cell].font = HDR
    ws[cell].fill = HDR_FILL


def inp(ws, cell, value, fmt=None):
    ws[cell] = value
    ws[cell].font = INPUT_FONT
    ws[cell].fill = INPUT_FILL
    ws[cell].border = BOX
    if fmt:
        ws[cell].number_format = fmt


# ---- Assumptions tab -------------------------------------------------------
def build_assumptions(ws, dcf):
    w = dcf.build_wacc()
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    for c in "CDEF":
        ws.column_dimensions[c].width = 13

    style_title(ws, "B2", "Tap & Turf Holdings Pty Ltd — DCF Assumptions")
    ws["B3"] = "Project Boundary | Round 1 Indicative Bid | AUD millions"
    ws["B3"].font = Font(italic=True, color="808080")

    ws["B5"] = "Last historical FYE"; ws["B5"].font = BOLD; ws["C5"] = "FY25 (Mar-25)"
    ws["B6"] = "Transaction date";   ws["B6"].font = BOLD; ws["C6"] = "FY26 (Mar-26)"

    hdr(ws, "B8", "Forecast drivers"); hdr(ws, "C8", "Value")
    drivers = [
        ("Revenue growth (p.a.)", dcf.REV_GROWTH, PCT, "p_rev_growth"),
        ("EBITDA margin (flat)", dcf.EBITDA_MARGIN, PCT, "p_ebitda_margin"),
        ("D&A / revenue", dcf.DA_PCT_REV, PCT, "p_da_pct"),
        ("Capex / revenue", dcf.CAPEX_PCT_REV, PCT, "p_capex_pct"),
        ("Change in WC / revenue (outflow)", dcf.WC_PCT_REV, PCT, "p_wc_pct"),
        ("Tax rate", dcf.TAX_RATE, PCT, "p_tax_rate"),
        ("Perpetuity growth rate", dcf.PERP_GROWTH, PCT, "p_perp_g"),
    ]
    r = 9
    driver_rows = {}
    for label, val, fmt, name in drivers:
        ws[f"B{r}"] = label
        inp(ws, f"C{r}", val, fmt)
        ws[f"D{r}"] = name; ws[f"D{r}"].font = Font(italic=True, color="808080")
        driver_rows[name] = r
        r += 1
    # WACC sits just below, computed from the build further down
    ws[f"B{r}"] = "WACC (= cost of equity, all-equity)"; ws[f"B{r}"].font = BOLD
    wacc_row = r
    driver_rows["p_wacc"] = r
    r += 2

    # WACC build
    hdr(ws, f"B{r}", "WACC build — CAPM (all-equity: net cash, asset-light)")
    ws.merge_cells(f"B{r}:F{r}")
    r += 1
    for col, txt in zip("BCDE", ["Comparable", "Lev β", "D/E", "Unlev β"]):
        ws[f"{col}{r}"] = txt; ws[f"{col}{r}"].font = BOLD; ws[f"{col}{r}"].fill = SUBHDR_FILL
    r += 1
    comp_start = r
    for c in w["comps"]:
        ws[f"B{r}"] = c["name"]
        ws[f"C{r}"] = c["lev_beta"]; ws[f"C{r}"].number_format = "0.00"
        ws[f"D{r}"] = c["de"]; ws[f"D{r}"].number_format = PCT
        ws[f"E{r}"] = f"=C{r}/(1+(1-$C${driver_rows['p_tax_rate']})*D{r})"
        ws[f"E{r}"].number_format = "0.000"
        r += 1
    comp_end = r - 1
    ws[f"B{r}"] = "Average unlevered β"; ws[f"B{r}"].font = BOLD
    ws[f"E{r}"] = f"=AVERAGE(E{comp_start}:E{comp_end})"; ws[f"E{r}"].number_format = "0.000"; ws[f"E{r}"].font = BOLD
    beta_row = r
    r += 2

    rows = {}
    capm = [
        ("Risk-free rate (AU 10Y)", dcf.RISK_FREE, PCT, "rf"),
        ("Equity risk premium", dcf.ERP, PCT, "erp"),
        ("Specific risk premium", dcf.SPECIFIC_RISK, PCT, "srp"),
    ]
    for label, val, fmt, key in capm:
        ws[f"B{r}"] = label
        inp(ws, f"C{r}", val, fmt)
        rows[key] = r
        r += 1
    ws[f"B{r}"] = "Cost of equity (CAPM)"
    ws[f"C{r}"] = f"=C{rows['rf']}+E{beta_row}*C{rows['erp']}"
    ws[f"C{r}"].number_format = PCT
    ke_row = r
    r += 1
    ws[f"B{r}"] = "+ Specific risk premium → WACC"; ws[f"B{r}"].font = BOLD
    ws[f"C{r}"] = f"=C{ke_row}+C{rows['srp']}"
    ws[f"C{r}"].number_format = PCT; ws[f"C{r}"].font = BOLD
    wacc_calc_row = r

    # Wire the headline WACC driver cell to the build result
    ws[f"C{wacc_row}"] = f"=C{wacc_calc_row}"
    ws[f"C{wacc_row}"].number_format = PCT
    ws[f"C{wacc_row}"].font = BOLD

    ws[f"B{r+2}"] = ("Note: valued cash-free / debt-free (net cash, no debt) → all-equity WACC, "
                     "EV = equity value. No exit multiple (no comparable set in the case).")
    ws[f"B{r+2}"].font = Font(italic=True, color="808080")
    return driver_rows


# ---- Company fin forecasts tab --------------------------------------------
def build_company_forecasts(ws, dcf):
    A = "Assumptions!$C$"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    style_title(ws, "B2", "Company financial forecasts")
    ws["B3"] = "FY21–FY25 actuals · FY26–FY31 forecast (driven from Assumptions)"
    ws["B3"].font = Font(italic=True, color="808080")

    hist_years = ["FY21", "FY22", "FY23", "FY24", "FY25"]
    fc_years = ["FY26", "FY27", "FY28", "FY29", "FY30", "FY31"]
    all_years = hist_years + fc_years
    first_col = 3  # column C
    cols = [get_column_letter(first_col + i) for i in range(len(all_years))]
    for c in cols:
        ws.column_dimensions[c].width = 11
    hist_cols = cols[:5]
    fc_cols = cols[5:]

    # header row 5
    hr = 5
    ws[f"B{hr}"] = "$m"; ws[f"B{hr}"].font = BOLD
    for col, yr in zip(cols, all_years):
        ws[f"{col}{hr}"] = yr
        ws[f"{col}{hr}"].font = HDR
        ws[f"{col}{hr}"].fill = HDR_FILL
        ws[f"{col}{hr}"].alignment = Alignment(horizontal="center")
    # mark forecast boundary
    ws[f"{fc_cols[0]}{hr-1}"] = "Forecast →"
    ws[f"{fc_cols[0]}{hr-1}"].font = Font(italic=True, bold=True, color="C00000")

    H = dcf.HISTORICAL

    def put_hist(row, key, sign=1):
        for col, yr in zip(hist_cols, hist_years):
            ws[f"{col}{row}"] = sign * H[yr][key]
            ws[f"{col}{row}"].number_format = MONEY

    rows = {}
    r = hr + 1

    def label(row, text, bold=False):
        ws[f"B{row}"] = text
        if bold:
            ws[f"B{row}"].font = BOLD

    # Revenue
    label(r, "Revenue", True); rows["rev"] = r
    put_hist(r, "rev")
    prev = hist_cols[-1]
    for col in fc_cols:
        ws[f"{col}{r}"] = f"={prev}{r}*(1+{A}9)"  # growth driver row 9
        ws[f"{col}{r}"].number_format = MONEY
        prev = col
    r += 1
    # growth %
    label(r, "  YoY growth %")
    for i, col in enumerate(cols):
        if i == 0:
            continue
        ws[f"{col}{r}"] = f"={col}{rows['rev']}/{cols[i-1]}{rows['rev']}-1"
        ws[f"{col}{r}"].number_format = PCT
    r += 1
    # EBITDA
    label(r, "EBITDA", True); rows["ebitda"] = r
    put_hist(r, "ebitda_clean")
    for col in fc_cols:
        ws[f"{col}{r}"] = f"={col}{rows['rev']}*{A}10"  # margin row 10
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    label(r, "  EBITDA margin %")
    for col in cols:
        ws[f"{col}{r}"] = f"={col}{rows['ebitda']}/{col}{rows['rev']}"
        ws[f"{col}{r}"].number_format = PCT
    r += 1
    # D&A
    label(r, "Depreciation & amortisation"); rows["da"] = r
    put_hist(r, "da", sign=-1)
    for col in fc_cols:
        ws[f"{col}{r}"] = f"=-{col}{rows['rev']}*{A}11"  # D&A% row 11
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    # EBIT
    label(r, "EBIT", True); rows["ebit"] = r
    for col in cols:
        ws[f"{col}{r}"] = f"={col}{rows['ebitda']}+{col}{rows['da']}"
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    # Tax
    label(r, "Tax on EBIT"); rows["tax"] = r
    for col in cols:
        ws[f"{col}{r}"] = f"=-{col}{rows['ebit']}*{A}14"  # tax rate row 14
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    # Capex
    label(r, "Capex"); rows["capex"] = r
    put_hist(r, "capex", sign=-1)
    for col in fc_cols:
        ws[f"{col}{r}"] = f"=-{col}{rows['rev']}*{A}12"  # capex% row 12
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    # WC
    label(r, "Change in working capital"); rows["wc"] = r
    for col in hist_cols:
        ws[f"{col}{r}"] = 0
        ws[f"{col}{r}"].number_format = MONEY
    for col in fc_cols:
        ws[f"{col}{r}"] = f"=-{col}{rows['rev']}*{A}13"  # WC% row 13
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    # Dividend
    label(r, "Dividend"); rows["div"] = r
    for col in cols:
        ws[f"{col}{r}"] = -24
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    # Net cash flow (post-dividend)
    label(r, "Net cash flow (post-dividend)", True); rows["ncf"] = r
    for col in cols:
        ws[f"{col}{r}"] = (f"={col}{rows['ebitda']}+{col}{rows['tax']}+{col}{rows['capex']}"
                           f"+{col}{rows['wc']}+{col}{rows['div']}")
        ws[f"{col}{r}"].number_format = MONEY
    r += 2
    return rows


# ---- DCF tab ---------------------------------------------------------------
def build_dcf(ws, dcf, cff_rows):
    A = "Assumptions!$C$"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    style_title(ws, "B2", "DCF — Perpetuity Growth (primary method)")
    ws["B3"] = "Unlevered FCF, mid-year discounting. AUD millions."
    ws["B3"].font = Font(italic=True, color="808080")

    fc_years = ["FY26", "FY27", "FY28", "FY29", "FY30", "FY31"]
    cols = [get_column_letter(3 + i) for i in range(len(fc_years))]
    for c in cols:
        ws.column_dimensions[c].width = 11
    # link to Company fin forecasts forecast columns (H..M there = fc cols)
    cff_fc_cols = [get_column_letter(3 + 5 + i) for i in range(6)]  # H..M

    hr = 5
    ws[f"B{hr}"] = "$m"; ws[f"B{hr}"].font = BOLD
    for col, yr in zip(cols, fc_years):
        ws[f"{col}{hr}"] = yr
        ws[f"{col}{hr}"].font = HDR; ws[f"{col}{hr}"].fill = HDR_FILL
        ws[f"{col}{hr}"].alignment = Alignment(horizontal="center")

    r = hr + 1
    rows = {}

    def line(text, bold=False):
        ws[f"B{r}"] = text
        if bold:
            ws[f"B{r}"].font = BOLD

    line("EBIT"); rows["ebit"] = r
    for col, cc in zip(cols, cff_fc_cols):
        ws[f"{col}{r}"] = f"='Company fin forecasts'!{cc}{cff_rows['ebit']}"
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    line("Less: tax on EBIT"); rows["tax"] = r
    for col in cols:
        ws[f"{col}{r}"] = f"=-{col}{rows['ebit']}*{A}14"
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    line("NOPAT"); rows["nopat"] = r
    for col in cols:
        ws[f"{col}{r}"] = f"={col}{rows['ebit']}+{col}{rows['tax']}"
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    line("Add: D&A"); rows["da"] = r
    for col, cc in zip(cols, cff_fc_cols):
        ws[f"{col}{r}"] = f"=-'Company fin forecasts'!{cc}{cff_rows['da']}"
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    line("Less: capex"); rows["capex"] = r
    for col, cc in zip(cols, cff_fc_cols):
        ws[f"{col}{r}"] = f"='Company fin forecasts'!{cc}{cff_rows['capex']}"
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    line("Less: change in WC"); rows["wc"] = r
    for col, cc in zip(cols, cff_fc_cols):
        ws[f"{col}{r}"] = f"='Company fin forecasts'!{cc}{cff_rows['wc']}"
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    line("Unlevered FCF", True); rows["fcf"] = r
    for col in cols:
        ws[f"{col}{r}"] = (f"={col}{rows['nopat']}+{col}{rows['da']}+{col}{rows['capex']}"
                           f"+{col}{rows['wc']}")
        ws[f"{col}{r}"].number_format = MONEY
    r += 1
    line("Discount period (yrs, mid-year)")
    rows["t"] = r
    for i, col in enumerate(cols):
        ws[f"{col}{r}"] = i + 0.5
        ws[f"{col}{r}"].number_format = "0.0"
    r += 1
    line("Discount factor @ WACC")
    rows["df"] = r
    for col in cols:
        ws[f"{col}{r}"] = f"=1/(1+{A}{15})^{col}{rows['t']}"  # WACC row 15 on Assumptions
        ws[f"{col}{r}"].number_format = "0.000"
    r += 1
    line("PV of FCF", True); rows["pv"] = r
    for col in cols:
        ws[f"{col}{r}"] = f"={col}{rows['fcf']}*{col}{rows['df']}"
        ws[f"{col}{r}"].number_format = MONEY
    r += 2

    last = cols[-1]
    # Valuation block
    hdr(ws, "B" + str(r), "Valuation (perpetuity growth)")
    ws.merge_cells(f"B{r}:C{r}")
    r += 1
    ws[f"B{r}"] = "Sum of PV(FCF)"
    ws[f"C{r}"] = f"=SUM({cols[0]}{rows['pv']}:{cols[-1]}{rows['pv']})"
    ws[f"C{r}"].number_format = MONEY
    sum_pv = r
    r += 1
    ws[f"B{r}"] = "Terminal value (Gordon, undiscounted)"
    ws[f"C{r}"] = f"={last}{rows['fcf']}*(1+{A}{ 'undef' })"  # placeholder, fixed below
    tv_row = r
    # proper TV formula: FCF_last*(1+g)/(WACC-g); g = Assumptions C15 (perp growth row 15? careful)
    r += 1
    ws[f"B{r}"] = "PV of terminal value"
    pv_tv_row = r
    r += 1
    ws[f"B{r}"] = "Enterprise value"; ws[f"B{r}"].font = BOLD
    ev_row = r
    r += 1
    ws[f"B{r}"] = "Net debt (cash-free/debt-free)"
    ws[f"C{r}"] = 0; ws[f"C{r}"].number_format = MONEY
    nd_row = r
    r += 1
    ws[f"B{r}"] = "Equity value (= EV)"; ws[f"B{r}"].font = BOLD
    eq_row = r
    r += 2
    ws[f"B{r}"] = "Check: implied EV / LFY EBITDA"
    chk_row = r

    return {"rows": rows, "cols": cols, "last": last, "sum_pv": sum_pv, "tv": tv_row,
            "pv_tv": pv_tv_row, "ev": ev_row, "nd": nd_row, "eq": eq_row, "chk": chk_row}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dst", required=True)
    args = ap.parse_args()
    dcf = load_dcf()

    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Assumptions"
    driver_rows = build_assumptions(ws_a, dcf)

    ws_c = wb.create_sheet("Company fin forecasts")
    cff_rows = build_company_forecasts(ws_c, dcf)

    ws_d = wb.create_sheet("DCF")
    d = build_dcf(ws_d, dcf, cff_rows)

    # Resolve Assumptions row references now that we know them
    g_row = driver_rows["p_perp_g"]
    wacc_row = driver_rows["p_wacc"]
    tax_row = driver_rows["p_tax_rate"]

    # Fix DCF formulas that needed assumption rows
    last = d["last"]
    fcf_r = d["rows"]["fcf"]
    # tax line uses WACC row placeholder C14 -> should be tax row
    for col in d["cols"]:
        ws_d[f"{col}{d['rows']['tax']}"] = f"=-{col}{d['rows']['ebit']}*Assumptions!$C${tax_row}"
        ws_d[f"{col}{d['rows']['df']}"] = f"=1/(1+Assumptions!$C${wacc_row})^{col}{d['rows']['t']}"
    # TV, PV(TV), EV, equity, check
    ws_d[f"C{d['tv']}"] = (f"={last}{fcf_r}*(1+Assumptions!$C${g_row})"
                           f"/(Assumptions!$C${wacc_row}-Assumptions!$C${g_row})")
    ws_d[f"C{d['tv']}"].number_format = MONEY
    ws_d[f"C{d['pv_tv']}"] = f"=C{d['tv']}/(1+Assumptions!$C${wacc_row})^{len(d['cols'])}"
    ws_d[f"C{d['pv_tv']}"].number_format = MONEY
    ws_d[f"C{d['ev']}"] = f"=C{d['sum_pv']}+C{d['pv_tv']}"
    ws_d[f"C{d['ev']}"].number_format = MONEY; ws_d[f"C{d['ev']}"].font = BOLD
    ws_d[f"C{d['eq']}"] = f"=C{d['ev']}-C{d['nd']}"
    ws_d[f"C{d['eq']}"].number_format = MONEY; ws_d[f"C{d['eq']}"].font = BOLD
    # implied EV/LFY EBITDA — FY25 clean EBITDA
    ws_d[f"C{d['chk']}"] = f"=C{d['ev']}/{dcf.HISTORICAL['FY25']['ebitda_clean']}"
    ws_d[f"C{d['chk']}"].number_format = MULT

    pathlib.Path(args.dst).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.dst)
    print(f"Wrote {args.dst}")


if __name__ == "__main__":
    main()
