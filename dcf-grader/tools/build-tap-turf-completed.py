"""Open the Tap & Turf template xlsx and write in completed numbers.

Source values are computed in tap-turf-dcf.py. We write VALUES (not
formulas) into each input cell — the template formulas downstream
recalculate when opened in Excel.

Run from repo root:
    python3 dcf-grader/tools/build-tap-turf-completed.py \\
        --src "/path/to/Tap_Turf_Corrupt.xlsx" \\
        --dst "dcf-grader/sample-models/tap-turf-completed.xlsx"
"""

import argparse
import importlib.util
import pathlib
import re
import shutil
import sys
import tempfile
import zipfile

from openpyxl import load_workbook


def sanitize_source(src: str) -> str:
    """Return a path to a copy of `src` with all external links removed.

    The template carries ~25 FactSet/Bloomberg add-in links to workbooks that
    don't exist. openpyxl round-trips them inconsistently (workbook.xml ends up
    declaring more <externalReference>s than there are link parts), which makes
    Excel refuse to open the file or demand repair. None of them are needed for
    the case study, so we strip the parts and every reference to them.
    """
    tmp = tempfile.mkdtemp(prefix="ttx_")
    extracted = pathlib.Path(tmp) / "x"
    with zipfile.ZipFile(src) as z:
        z.extractall(extracted)

    # 1. Drop the externalLinks directory and calcChain (Excel rebuilds calcChain).
    shutil.rmtree(extracted / "xl" / "externalLinks", ignore_errors=True)
    (extracted / "xl" / "calcChain.xml").unlink(missing_ok=True)

    # 2. Strip <externalReferences>…</externalReferences> from workbook.xml, plus any
    #    defined name that points at an external workbook ([1], [2], …) — those would
    #    dangle as #REF! once the links are gone.
    wb_xml = extracted / "xl" / "workbook.xml"
    text = wb_xml.read_text(encoding="utf-8")
    text = re.sub(r"<externalReferences>.*?</externalReferences>", "", text, flags=re.DOTALL)
    text = re.sub(r"<definedName\b[^>]*>[^<]*\[\d+\][^<]*</definedName>", "", text)
    wb_xml.write_text(text, encoding="utf-8")

    # 3. Strip externalLink + calcChain relationships from workbook rels.
    rels = extracted / "xl" / "_rels" / "workbook.xml.rels"
    rtext = rels.read_text(encoding="utf-8")
    rtext = re.sub(r'<Relationship[^>]*(externalLink|calcChain)[^>]*/>', "", rtext)
    rels.write_text(rtext, encoding="utf-8")

    # 4. Strip externalLink + calcChain overrides from [Content_Types].xml.
    ct = extracted / "[Content_Types].xml"
    ctext = ct.read_text(encoding="utf-8")
    ctext = re.sub(r'<Override[^>]*(externalLink|calcChain)[^>]*/>', "", ctext)
    ct.write_text(ctext, encoding="utf-8")

    clean = pathlib.Path(tmp) / "clean.xlsx"
    with zipfile.ZipFile(clean, "w", zipfile.ZIP_DEFLATED) as z:
        for p in sorted(extracted.rglob("*")):
            if p.is_file():
                z.write(p, p.relative_to(extracted).as_posix())
    return str(clean)


def load_dcf_module():
    here = pathlib.Path(__file__).resolve().parent
    spec = importlib.util.spec_from_file_location("tap_turf_dcf", here / "tap-turf-dcf.py")
    module = importlib.util.module_from_spec(spec)
    sys.modules["tap_turf_dcf"] = module
    spec.loader.exec_module(module)
    return module


def write(ws, addr: str, value) -> None:
    ws[addr] = value


def fill_financials_tab(ws, hist, fc) -> None:
    """Financials tab: populate L–P (FY21–FY25 historicals) and Q–V (FY26–FY31 forecast)."""
    years_hist = ["FY21", "FY22", "FY23", "FY24", "FY25"]
    cols_hist = ["L", "M", "N", "O", "P"]
    cols_fc   = ["Q", "R", "S", "T", "U", "V"]

    # Row 11 Total revenue: write the actual revenue figure (overrides L11=K11*(1+L13) formula)
    # Row 12 Management case revenue (input row)
    # Row 15 EBITDA (display) and Row 16 management case EBITDA (input)
    for col, yr in zip(cols_hist, years_hist):
        write(ws, f"{col}11", hist[yr]["rev"])
        write(ws, f"{col}12", hist[yr]["rev"])
        write(ws, f"{col}15", hist[yr]["ebitda_clean"])
        write(ws, f"{col}16", hist[yr]["ebitda_clean"])
        write(ws, f"{col}18", hist[yr]["ebitda_clean"] / hist[yr]["rev"])
        write(ws, f"{col}20", -hist[yr]["da"])
        write(ws, f"{col}25", hist[yr]["ebitda_clean"] - hist[yr]["da"])  # EBIT
        write(ws, f"{col}29", 0.19)                                        # tax rate
        write(ws, f"{col}40", -hist[yr]["capex"])
        # change in WC — case-study says ~ -1.5% of revenue is the forward assumption;
        # historical figures aren't directly given as ΔNWC, so we use -1.5% × rev as approximation
        write(ws, f"{col}45", -hist[yr]["rev"] * 0.015)
        write(ws, f"{col}47", -0.015)
        write(ws, f"{col}49", 0)
        write(ws, f"{col}53", 0)

    # Forecast columns Q-V: all driver inputs live on the Assumptions tab (D27:D35).
    # Cells here are pure formulas — change a driver on Assumptions, whole model recalcs.
    A = "Assumptions!$D$"   # shorthand for absolute refs to the assumptions block
    prev_col = "P"
    for col in cols_fc:
        write(ws, f"{col}13", f"={A}27")                              # growth (link)
        write(ws, f"{col}18", f"={A}28")                              # margin (link)
        write(ws, f"{col}29", f"={A}32")                              # tax rate (link)
        write(ws, f"{col}47", f"=-{A}31")                             # WC % (link, negative)

        write(ws, f"{col}11", f"={prev_col}11*(1+{col}13)")           # Revenue = prior × (1+g)
        write(ws, f"{col}12", f"={col}11")                            # Management case
        write(ws, f"{col}15", f"={col}11*{col}18")                    # EBITDA = revenue × margin
        write(ws, f"{col}16", f"={col}15")
        write(ws, f"{col}20", f"=-{col}11*{A}29")                     # D&A = -revenue × D&A%
        write(ws, f"{col}25", f"={col}15+{col}20")                    # EBIT
        write(ws, f"{col}31", f"={col}25*(1-{col}29)")                # Net income (~EBIT × (1–t))
        write(ws, f"{col}40", f"=-{col}11*{A}30")                     # Capex = -revenue × capex%
        write(ws, f"{col}42", f"={A}30")
        write(ws, f"{col}45", f"={col}47*{col}11")                    # ΔWC = WC% × revenue
        write(ws, f"{col}49", 0)                                      # Other CF
        write(ws, f"{col}53", 0)                                      # Exceptional
        prev_col = col


def fill_dcf_input_tab(ws, hist, fc, params) -> None:
    """DCF input tab: assumptions block + FCF build (all forecast cells link to Assumptions)."""
    A = "Assumptions!$D$"
    write(ws, "F9", f"={A}34")     # perp growth → Assumptions D34
    write(ws, "F10", f"={A}35")    # exit multiple → Assumptions D35
    write(ws, "F11", f"={A}33")    # WACC → Assumptions D33

    # Column F=FY18, G=FY19, H=FY20, I=FY21, J=FY22, K=FY23, L=FY24, M=FY25, N=FY26..S=FY31
    historicals = {"I": "FY21", "J": "FY22", "K": "FY23", "L": "FY24", "M": "FY25"}
    for col, yr in historicals.items():
        h = hist[yr]
        write(ws, f"{col}24", h["rev"])
        write(ws, f"{col}25", 0)  # leave growth blank for historicals — formulas compute it
        write(ws, f"{col}27", h["ebitda_clean"])
        write(ws, f"{col}28", h["ebitda_clean"] / h["rev"])
        write(ws, f"{col}31", -h["da"])
        write(ws, f"{col}32", h["da"] / h["rev"])
        write(ws, f"{col}33", h["da"] / h["capex"])
        write(ws, f"{col}35", h["ebitda_clean"] - h["da"])
        write(ws, f"{col}36", (h["ebitda_clean"] - h["da"]) / h["rev"])
        write(ws, f"{col}39", -(h["ebitda_clean"] - h["da"]) * 0.19)
        write(ws, f"{col}40", 0.19)
        write(ws, f"{col}42", (h["ebitda_clean"] - h["da"]) * (1 - 0.19))
        write(ws, f"{col}46", h["da"])
        write(ws, f"{col}48", -h["capex"])
        write(ws, f"{col}49", -h["capex"] / h["rev"])
        write(ws, f"{col}50", h["da"] / h["capex"])
        write(ws, f"{col}52", -h["rev"] * 0.015)
        write(ws, f"{col}53", -0.015)
        write(ws, f"{col}55", 0)
        write(ws, f"{col}56", 0)
        write(ws, f"{col}58", 0)
        write(ws, f"{col}59", 0)

    forecast_cols = ["N", "O", "P", "Q", "R", "S"]
    prev_col = "M"
    for col in forecast_cols:
        # Driver inputs link to Assumptions
        write(ws, f"{col}25", f"={A}27")     # revenue growth
        write(ws, f"{col}28", f"={A}28")     # EBITDA margin
        write(ws, f"{col}40", f"={A}32")     # tax rate
        write(ws, f"{col}49", f"=-{A}30")    # capex % of revenue (negative display)
        write(ws, f"{col}53", f"=-{A}31")    # WC % of revenue (negative display)
        # Formula-driven
        write(ws, f"{col}24", f"={prev_col}24*(1+{col}25)")            # Revenue
        write(ws, f"{col}27", f"={col}24*{col}28")                     # EBITDA
        write(ws, f"{col}31", f"=-{col}24*{A}29")                      # D&A
        write(ws, f"{col}32", f"={A}29")                               # D&A / revenue
        write(ws, f"{col}35", f"={col}27+{col}31")                     # EBIT
        write(ws, f"{col}36", f"={col}35/{col}24")                     # EBIT margin
        write(ws, f"{col}39", f"=-{col}35*{col}40")                    # Tax
        write(ws, f"{col}42", f"={col}35+{col}39")                     # EBIAT
        write(ws, f"{col}43", f"={col}42/{col}24")                     # EBIAT margin
        write(ws, f"{col}46", f"=-{col}31")                            # D&A addback
        write(ws, f"{col}48", f"={col}49*{col}24")                     # Capex
        write(ws, f"{col}50", f"=-{col}46/{col}48")                    # D&A / capex
        write(ws, f"{col}52", f"={col}53*{col}24")                     # ΔWC
        write(ws, f"{col}55", 0)
        write(ws, f"{col}56", 0)
        write(ws, f"{col}58", 0)
        write(ws, f"{col}59", 0)
        write(ws, f"{col}61", f"={col}42+{col}46+{col}48+{col}52+{col}55+{col}58")
        write(ws, f"{col}64", f"={col}63+{col}61")
        prev_col = col

    # Summary block — Row 111 dates, then summary rows 112-127 = same as above. The template
    # has formulas referencing F24/F27 etc. — those will recalc on open. We don't need to write.


def fill_dcf_output_tab(ws, fc, valuation, params) -> None:
    """DCF output tab: explicit forecast period (FY26-FY31) and TV column T."""
    forecast_cols = ["I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]
    # The DCF output uses I=FY21, J=FY22, K=FY23, L=FY24, M=FY25, N=FY26..S=FY31
    # Forecast years are N-S
    historicals = {"I": 0, "J": 1, "K": 2, "L": 3, "M": 4}
    forecast_map = {"N": 0, "O": 1, "P": 2, "Q": 3, "R": 4, "S": 5}

    # Historical block in DCF output (rows 9-28) — minimal: revenue, EBITDA, etc.
    hist_data = [
        (1149.23, 94.15, 36.13, 36.0),
        (1255.63, 92.62, 40.46, 51.0),
        (1353.84, 111.80, 47.20, 50.0),
        (1446.98, 124.40, 51.83, 50.0),
        (1442.62, 127.57, 54.80, 50.0),
    ]
    for col, idx in historicals.items():
        rev, ebitda, da, capex = hist_data[idx]
        write(ws, f"{col}9", rev)
        write(ws, f"{col}11", ebitda)
        write(ws, f"{col}14", -da)
        write(ws, f"{col}17", ebitda - da)
        write(ws, f"{col}21", -capex)

    for col, idx in forecast_map.items():
        r = fc[idx]
        write(ws, f"{col}9", r["rev"])
        write(ws, f"{col}11", r["ebitda"])
        write(ws, f"{col}14", -r["da"])
        write(ws, f"{col}17", r["ebit"])
        write(ws, f"{col}19", -r["tax"])
        write(ws, f"{col}20", 0.19)
        write(ws, f"{col}21", -r["capex"])
        write(ws, f"{col}23", -r["wc"])
        write(ws, f"{col}24", 0)
        write(ws, f"{col}26", 0)
        write(ws, f"{col}28", r["fcf"])

    # TV column T (perpetuity method)
    write(ws, "T28", valuation["tv_perp"])

    # Summary valuation
    write(ws, "F37", valuation["pv_fcf"])
    write(ws, "F38", valuation["pv_tv_perp"])
    write(ws, "F39", valuation["ev_perp"])
    # F40 net debt already in template


def fill_assumptions_tab(ws, params) -> None:
    """Assumptions tab: add a Forecast drivers block so a single cell controls everything.

    Rows 26-34 are unused in the template — putting the driver inputs here so the
    candidate (or Archie) can change the 6% growth in one place and have the whole
    model recalculate.
    """
    write(ws, "C26", "Forecast drivers")
    write(ws, "C27", "Revenue growth (annual)")
    write(ws, "D27", 0.06)
    write(ws, "E27", "p_rev_growth")
    write(ws, "C28", "EBITDA margin (flat)")
    write(ws, "D28", 0.0872)
    write(ws, "E28", "p_ebitda_margin")
    write(ws, "C29", "D&A / revenue")
    write(ws, "D29", 0.038)
    write(ws, "E29", "p_da_pct")
    write(ws, "C30", "Capex / revenue")
    write(ws, "D30", 0.03)
    write(ws, "E30", "p_capex_pct")
    write(ws, "C31", "ΔWC / revenue (outflow)")
    write(ws, "D31", 0.015)
    write(ws, "E31", "p_wc_pct")
    write(ws, "C32", "Effective tax rate")
    write(ws, "D32", 0.19)
    write(ws, "E32", "p_tax_rate")
    write(ws, "C33", "WACC")
    write(ws, "D33", params["wacc"])
    write(ws, "E33", "p_wacc")
    write(ws, "C34", "Perpetuity growth rate")
    write(ws, "D34", params["perp_growth"])
    write(ws, "E34", "p_perp_g")
    write(ws, "C35", "Exit EBITDA multiple (peer median)")
    write(ws, "D35", "=D43")                          # link to comps median below
    write(ws, "E35", "p_exit_multiple")

    # Comparable trading multiples — same peer set as the WACC beta build.
    write(ws, "C37", "Comparable trading multiples")
    write(ws, "D37", "EV/EBITDA")
    comps = params.get("comps", [])
    r = 38
    for c in comps:
        write(ws, f"C{r}", c["name"])
        write(ws, f"D{r}", c["ev_ebitda"])
        r += 1
    write(ws, "C43", "Median EV/EBITDA")
    write(ws, "D43", f"=MEDIAN(D38:D{r-1})")

    # Cash-free / debt-free deal basis → no net debt adjustment in the equity bridge.
    write(ws, "D19", 0)


def fill_company_fin_forecasts_tab(ws) -> None:
    """Extend the 5-year plan with FY26-FY31 and strip ALL debt (per no-debt structure).

    The business is all-equity (net cash, no facilities), so every interest line
    (rows 12/17/26) and the net-debt roll (rows 34-36) are zeroed across all years.
    EBITDA is set directly (historical = reported post-exceptional; forecast =
    margin × revenue) and the P&L is re-derived from it, so removing interest does
    NOT shift EBITDA/EBIT. Dividends held at $24m/yr; WC is a forecast outflow.
    """
    A = "Assumptions!$D$"
    hist_cols = ["B", "C", "D", "E", "F"]
    hist_ebitda = {"B": 94.15, "C": 74.62, "D": 105.04, "E": 125.53, "F": 122.70}
    fc_cols = ["G", "H", "I", "J", "K", "L"]
    fy_labels = ["FY26", "FY27", "FY28", "FY29", "FY30", "FY31"]
    all_cols = hist_cols + fc_cols

    write(ws, "G3", "Forecast (assumption-driven, FY26+) →")

    # Forecast P&L / cashflow — no interest, no debt
    prev = "F"
    for col, label in zip(fc_cols, fy_labels):
        write(ws, f"{col}4", label)
        write(ws, f"{col}5", f"={prev}5*(1+{A}27)")          # Revenue ×(1+g)
        write(ws, f"{col}6", f"={col}5/{prev}5-1")           # YoY growth
        write(ws, f"{col}7", f"={col}5*$F$8")                # GP = flat FY25 GM% × rev
        write(ws, f"{col}8", f"={col}7/{col}5")              # GM%
        write(ws, f"{col}10", f"={col}16-{col}15")           # Operating profit = EBIT (no interest)
        write(ws, f"{col}9", f"={col}10-{col}7")             # Opex (plug)
        write(ws, f"{col}11", f"={col}10/{col}5")            # Operating margin
        write(ws, f"{col}15", 0)                             # Exceptionals (nil)
        write(ws, f"{col}18", "=$F$18")                      # Amortisation (flat)
        write(ws, f"{col}19", f"={col}5*{A}29-{col}18")      # Depreciation = D&A − amort
        write(ws, f"{col}20", f"={col}5*{A}28")              # EBITDA = margin × rev (DRIVER)
        write(ws, f"{col}16", f"={col}20-{col}18-{col}19")   # EBIT (= PBT, no interest)
        write(ws, f"{col}13", f"={col}16-{col}15")           # PBT underlying
        write(ws, f"{col}22", 0)                             # Non-cash exc adjust
        write(ws, f"{col}23", f"=-{col}5*{A}30")             # Capex = −3% × rev
        write(ws, f"{col}24", 0)                             # Acquisition
        write(ws, f"{col}25", f"=-{col}16*{A}32")            # Tax = −EBIT × tax
        write(ws, f"{col}27", f"=-{col}5*{A}31")             # Working capital OUTFLOW (−1.5%)
        write(ws, f"{col}28", "=$F$28")                      # Dividend (−$24m, held)
        write(ws, f"{col}29", 0)                             # Other (nil)
        write(ws, f"{col}31", f"=SUM({col}20:{col}30)")      # Net cash flow
        write(ws, f"{col}32", f"={col}31-{col}28-{col}24")   # Free cash flow
        prev = col

    # Historical: re-derive EBIT/PBT from reported EBITDA so zeroing interest
    # doesn't change the reported EBITDA.
    for col in hist_cols:
        write(ws, f"{col}20", hist_ebitda[col])              # EBITDA (reported, post-exc)
        write(ws, f"{col}16", f"={col}20-{col}18-{col}19")   # EBIT (no interest)
        write(ws, f"{col}13", f"={col}16-{col}15")           # PBT underlying

    # Strip every debt / interest line across ALL years (historical + forecast)
    for col in all_cols:
        write(ws, f"{col}12", 0)   # Interest and Amortisation
        write(ws, f"{col}17", 0)   # Add back: Interest
        write(ws, f"{col}26", 0)   # Interest (cash)
        write(ws, f"{col}34", 0)   # Opening Net Debt
        write(ws, f"{col}35", 0)   # Closing Net Debt
        write(ws, f"{col}36", 0)   # Net debt : EBITDA



def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, help="path to Archie's template xlsx")
    ap.add_argument("--dst", required=True, help="output path for completed xlsx")
    args = ap.parse_args()

    dcf = load_dcf_module()
    fc = dcf.build_forecast()
    valuation = dcf.discount(fc)
    params = {
        "wacc": dcf.WACC,
        "perp_growth": dcf.PERP_GROWTH,
        # No comparable transactions/trading comps for stadium concessions, so we don't
        # assert an exit multiple. We store the perpetuity-IMPLIED multiple as a memo so
        # the template's exit-multiple block reconciles to perpetuity as a sanity check.
        "exit_multiple": dcf.EXIT_MULTIPLE,
        "comps": dcf.COMPS,
    }

    wb = load_workbook(sanitize_source(args.src))
    print(f"Sheets: {wb.sheetnames}")

    fill_assumptions_tab(wb["Assumptions"], params)
    fill_company_fin_forecasts_tab(wb["Company fin forecasts"])
    fill_financials_tab(wb[" Financials"], dcf.HISTORICAL, fc)
    fill_dcf_input_tab(wb["DCF input"], dcf.HISTORICAL, fc, params)
    fill_dcf_output_tab(wb["DCF output"], fc, valuation, params)

    pathlib.Path(args.dst).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.dst)
    print(f"Wrote {args.dst}")


if __name__ == "__main__":
    main()
