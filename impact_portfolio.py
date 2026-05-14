"""
Impact Investing Portfolio
Tracks holdings with financial returns, ESG scores, and tangible impact metrics.
"""

import csv
import json
import os
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


@dataclass
class Investment:
    """A single holding in the portfolio."""

    ticker: str
    name: str
    sector: str
    shares: float
    cost_basis: float       # USD per share at purchase
    current_price: float    # USD per share, latest

    # ESG scores: 0-100 (higher = better)
    e_score: float = 0.0
    s_score: float = 0.0
    g_score: float = 0.0

    # Per-share impact metrics (annual, attributable to one share)
    co2e_avoided_t: float = 0.0      # tonnes CO2-equivalent
    clean_energy_mwh: float = 0.0    # MWh of clean energy enabled
    people_served: float = 0.0       # people reached by products/services
    jobs_supported: float = 0.0      # jobs in underserved communities

    notes: str = ""

    def cost(self) -> float:
        return self.shares * self.cost_basis

    def value(self) -> float:
        return self.shares * self.current_price

    def gain(self) -> float:
        return self.value() - self.cost()

    def return_pct(self) -> float:
        c = self.cost()
        return (self.gain() / c * 100.0) if c else 0.0

    def esg_score(self) -> float:
        return (self.e_score + self.s_score + self.g_score) / 3.0

    def impact(self) -> Dict[str, float]:
        """Total impact for this holding (per-share metric * shares)."""
        return {
            "co2e_avoided_t": self.co2e_avoided_t * self.shares,
            "clean_energy_mwh": self.clean_energy_mwh * self.shares,
            "people_served": self.people_served * self.shares,
            "jobs_supported": self.jobs_supported * self.shares,
        }


class Portfolio:
    """A collection of impact investments with portfolio-level analytics."""

    IMPACT_KEYS = ("co2e_avoided_t", "clean_energy_mwh", "people_served", "jobs_supported")

    def __init__(self, name: str = "Impact Portfolio"):
        self.name = name
        self.holdings: List[Investment] = []

    # ---- Mutation -----------------------------------------------------------

    def add(self, investment: Investment) -> None:
        self.holdings.append(investment)

    def remove(self, ticker: str) -> bool:
        before = len(self.holdings)
        self.holdings = [h for h in self.holdings if h.ticker.upper() != ticker.upper()]
        return len(self.holdings) < before

    # ---- Financial analytics ------------------------------------------------

    def total_cost(self) -> float:
        return sum(h.cost() for h in self.holdings)

    def total_value(self) -> float:
        return sum(h.value() for h in self.holdings)

    def total_gain(self) -> float:
        return self.total_value() - self.total_cost()

    def total_return_pct(self) -> float:
        c = self.total_cost()
        return (self.total_gain() / c * 100.0) if c else 0.0

    def weights(self) -> Dict[str, float]:
        """Position weight by current market value."""
        v = self.total_value()
        if not v:
            return {h.ticker: 0.0 for h in self.holdings}
        return {h.ticker: h.value() / v for h in self.holdings}

    def sector_allocation(self) -> Dict[str, float]:
        """Sector weights by current market value."""
        v = self.total_value()
        out: Dict[str, float] = {}
        if not v:
            return out
        for h in self.holdings:
            out[h.sector] = out.get(h.sector, 0.0) + h.value() / v
        return out

    # ---- ESG analytics ------------------------------------------------------

    def weighted_esg(self) -> Dict[str, float]:
        """Value-weighted E, S, G, and combined score across the portfolio."""
        v = self.total_value()
        if not v:
            return {"E": 0.0, "S": 0.0, "G": 0.0, "ESG": 0.0}
        e = sum(h.e_score * h.value() for h in self.holdings) / v
        s = sum(h.s_score * h.value() for h in self.holdings) / v
        g = sum(h.g_score * h.value() for h in self.holdings) / v
        return {"E": e, "S": s, "G": g, "ESG": (e + s + g) / 3.0}

    # ---- Impact analytics ---------------------------------------------------

    def total_impact(self) -> Dict[str, float]:
        totals = {k: 0.0 for k in self.IMPACT_KEYS}
        for h in self.holdings:
            for k, val in h.impact().items():
                totals[k] += val
        return totals

    def impact_per_1k_invested(self) -> Dict[str, float]:
        """Impact normalized per $1,000 of cost basis — useful for comparing portfolios."""
        cost = self.total_cost()
        totals = self.total_impact()
        if not cost:
            return {k: 0.0 for k in totals}
        return {k: v / (cost / 1000.0) for k, v in totals.items()}

    # ---- I/O ----------------------------------------------------------------

    @classmethod
    def from_csv(cls, file_path: str, name: Optional[str] = None) -> "Portfolio":
        """Load holdings from a CSV file. See sample_portfolio.csv for the schema."""
        portfolio = cls(name=name or os.path.splitext(os.path.basename(file_path))[0])
        with open(file_path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                portfolio.add(_investment_from_row(row))
        return portfolio

    def to_dataframe(self) -> pd.DataFrame:
        rows = []
        for h in self.holdings:
            row = asdict(h)
            row["cost"] = round(h.cost(), 2)
            row["value"] = round(h.value(), 2)
            row["gain"] = round(h.gain(), 2)
            row["return_pct"] = round(h.return_pct(), 2)
            row["esg_score"] = round(h.esg_score(), 1)
            for k, v in h.impact().items():
                row[f"total_{k}"] = round(v, 2)
            rows.append(row)
        return pd.DataFrame(rows)

    def save_excel(self, file_path: str) -> str:
        """Write a formatted Excel report with Holdings + Summary sheets."""
        df = self.to_dataframe()
        summary = self._summary_dataframe()

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Holdings")
            summary.to_excel(writer, index=False, sheet_name="Summary")

        _format_workbook(file_path)
        return file_path

    def save_html(self, file_path: str) -> str:
        """Write a self-contained HTML dashboard (Chart.js via CDN)."""
        payload = {
            "name": self.name,
            "totals": {
                "cost": round(self.total_cost(), 2),
                "value": round(self.total_value(), 2),
                "gain": round(self.total_gain(), 2),
                "return_pct": round(self.total_return_pct(), 2),
            },
            "esg": {k: round(v, 1) for k, v in self.weighted_esg().items()},
            "sectors": {k: round(v * 100, 2) for k, v in self.sector_allocation().items()},
            "impact": {k: round(v, 2) for k, v in self.total_impact().items()},
            "per1k": {k: round(v, 3) for k, v in self.impact_per_1k_invested().items()},
            "holdings": [
                {
                    "ticker": h.ticker,
                    "name": h.name,
                    "sector": h.sector,
                    "shares": h.shares,
                    "cost": round(h.cost(), 2),
                    "value": round(h.value(), 2),
                    "gain": round(h.gain(), 2),
                    "return_pct": round(h.return_pct(), 2),
                    "esg": round(h.esg_score(), 1),
                }
                for h in self.holdings
            ],
        }
        html = _HTML_TEMPLATE.replace("__DATA__", json.dumps(payload))
        with open(file_path, "w") as f:
            f.write(html)
        return file_path

    # ---- Reporting ----------------------------------------------------------

    def summary(self) -> str:
        esg = self.weighted_esg()
        impact = self.total_impact()
        per1k = self.impact_per_1k_invested()
        sectors = self.sector_allocation()

        lines = []
        lines.append("=" * 64)
        lines.append(f"  {self.name}")
        lines.append("=" * 64)
        lines.append(f"  Holdings:        {len(self.holdings)}")
        lines.append(f"  Cost basis:      ${self.total_cost():,.2f}")
        lines.append(f"  Market value:    ${self.total_value():,.2f}")
        lines.append(
            f"  Unrealized P/L:  ${self.total_gain():,.2f} "
            f"({self.total_return_pct():+.2f}%)"
        )
        lines.append("")
        lines.append("  Weighted ESG (0-100):")
        lines.append(
            f"    E={esg['E']:.1f}  S={esg['S']:.1f}  G={esg['G']:.1f}  "
            f"ESG={esg['ESG']:.1f}"
        )
        lines.append("")
        lines.append("  Sector allocation:")
        for sector, w in sorted(sectors.items(), key=lambda kv: -kv[1]):
            lines.append(f"    {sector:<28} {w * 100:5.1f}%")
        lines.append("")
        lines.append("  Annual impact (totals):")
        lines.append(f"    CO2e avoided:        {impact['co2e_avoided_t']:>12,.1f}  tonnes")
        lines.append(f"    Clean energy:        {impact['clean_energy_mwh']:>12,.1f}  MWh")
        lines.append(f"    People served:       {impact['people_served']:>12,.0f}")
        lines.append(f"    Jobs supported:      {impact['jobs_supported']:>12,.0f}")
        lines.append("")
        lines.append("  Impact per $1,000 invested:")
        lines.append(f"    CO2e avoided:        {per1k['co2e_avoided_t']:>12,.3f}  tonnes")
        lines.append(f"    Clean energy:        {per1k['clean_energy_mwh']:>12,.3f}  MWh")
        lines.append(f"    People served:       {per1k['people_served']:>12,.2f}")
        lines.append(f"    Jobs supported:      {per1k['jobs_supported']:>12,.3f}")
        lines.append("=" * 64)
        return "\n".join(lines)

    def _summary_dataframe(self) -> pd.DataFrame:
        esg = self.weighted_esg()
        impact = self.total_impact()
        per1k = self.impact_per_1k_invested()
        rows = [
            ("Holdings", len(self.holdings)),
            ("Cost basis (USD)", round(self.total_cost(), 2)),
            ("Market value (USD)", round(self.total_value(), 2)),
            ("Unrealized P/L (USD)", round(self.total_gain(), 2)),
            ("Total return (%)", round(self.total_return_pct(), 2)),
            ("Weighted E", round(esg["E"], 1)),
            ("Weighted S", round(esg["S"], 1)),
            ("Weighted G", round(esg["G"], 1)),
            ("Weighted ESG", round(esg["ESG"], 1)),
            ("CO2e avoided (t/yr)", round(impact["co2e_avoided_t"], 1)),
            ("Clean energy (MWh/yr)", round(impact["clean_energy_mwh"], 1)),
            ("People served", round(impact["people_served"], 0)),
            ("Jobs supported", round(impact["jobs_supported"], 0)),
            ("CO2e per $1k", round(per1k["co2e_avoided_t"], 3)),
            ("MWh per $1k", round(per1k["clean_energy_mwh"], 3)),
            ("People per $1k", round(per1k["people_served"], 2)),
            ("Jobs per $1k", round(per1k["jobs_supported"], 3)),
        ]
        return pd.DataFrame(rows, columns=["Metric", "Value"])


# ---- Helpers ---------------------------------------------------------------


_NUMERIC_FIELDS = {
    "shares", "cost_basis", "current_price",
    "e_score", "s_score", "g_score",
    "co2e_avoided_t", "clean_energy_mwh", "people_served", "jobs_supported",
}


def _investment_from_row(row: Dict[str, str]) -> Investment:
    kwargs: Dict[str, object] = {}
    for key, val in row.items():
        key = key.strip()
        if key in _NUMERIC_FIELDS:
            kwargs[key] = float(val) if val not in (None, "") else 0.0
        else:
            kwargs[key] = (val or "").strip()
    return Investment(**kwargs)  # type: ignore[arg-type]


def _format_workbook(file_path: str) -> None:
    wb = load_workbook(file_path)

    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 32)

    wb.save(file_path)


# ---- Demo ------------------------------------------------------------------


def _demo_portfolio() -> Portfolio:
    """Build an in-memory sample portfolio for demonstration."""
    p = Portfolio("Sample Impact Portfolio")
    p.add(Investment(
        ticker="FSLR", name="First Solar", sector="Clean Energy",
        shares=40, cost_basis=160.00, current_price=205.50,
        e_score=88, s_score=72, g_score=78,
        co2e_avoided_t=1.20, clean_energy_mwh=4.5,
        notes="Utility-scale thin-film solar manufacturer",
    ))
    p.add(Investment(
        ticker="NEE", name="NextEra Energy", sector="Clean Energy",
        shares=120, cost_basis=68.40, current_price=74.10,
        e_score=82, s_score=70, g_score=80,
        co2e_avoided_t=0.85, clean_energy_mwh=2.1,
    ))
    p.add(Investment(
        ticker="BEPC", name="Brookfield Renewable", sector="Clean Energy",
        shares=90, cost_basis=29.80, current_price=27.55,
        e_score=86, s_score=68, g_score=74,
        co2e_avoided_t=0.95, clean_energy_mwh=3.0,
    ))
    p.add(Investment(
        ticker="BLDP", name="Ballard Power Systems", sector="Clean Mobility",
        shares=300, cost_basis=4.10, current_price=3.05,
        e_score=78, s_score=64, g_score=66,
        co2e_avoided_t=0.20,
    ))
    p.add(Investment(
        ticker="WM", name="Waste Management", sector="Circular Economy",
        shares=25, cost_basis=178.00, current_price=212.40,
        e_score=70, s_score=66, g_score=82,
        co2e_avoided_t=0.40,
    ))
    p.add(Investment(
        ticker="XYL", name="Xylem Inc.", sector="Water & Sanitation",
        shares=55, cost_basis=108.20, current_price=129.75,
        e_score=80, s_score=78, g_score=76,
        people_served=420,
    ))
    p.add(Investment(
        ticker="VEEV", name="Veeva Systems", sector="Health & Wellbeing",
        shares=35, cost_basis=185.00, current_price=234.10,
        e_score=64, s_score=82, g_score=80,
        people_served=180,
    ))
    p.add(Investment(
        ticker="GRMN", name="Garmin", sector="Health & Wellbeing",
        shares=20, cost_basis=120.00, current_price=147.30,
        e_score=66, s_score=72, g_score=78,
        people_served=95,
    ))
    p.add(Investment(
        ticker="SBUX", name="Starbucks", sector="Sustainable Food",
        shares=60, cost_basis=92.50, current_price=99.20,
        e_score=68, s_score=74, g_score=70,
        jobs_supported=0.08,
    ))
    p.add(Investment(
        ticker="JKHY", name="Jack Henry & Associates", sector="Financial Inclusion",
        shares=18, cost_basis=158.00, current_price=171.40,
        e_score=60, s_score=78, g_score=84,
        people_served=220, jobs_supported=0.05,
    ))
    return p


def main() -> None:
    print("=" * 64)
    print("  IMPACT INVESTING PORTFOLIO")
    print("=" * 64)
    print()

    sample_csv = "sample_portfolio.csv"
    if os.path.exists(sample_csv):
        print(f"Loading portfolio from {sample_csv} ...")
        portfolio = Portfolio.from_csv(sample_csv, name="Sample Impact Portfolio")
    else:
        print("No sample_portfolio.csv found — using built-in demo portfolio.")
        portfolio = _demo_portfolio()

    print()
    print(portfolio.summary())
    print()

    xlsx_out = "portfolio_report.xlsx"
    html_out = "portfolio_report.html"
    portfolio.save_excel(xlsx_out)
    portfolio.save_html(html_out)
    print(f"Wrote Excel report to:    {xlsx_out}")
    print(f"Wrote HTML dashboard to:  {html_out}")


_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Impact Investing Portfolio</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
  :root {
    --bg: #0f1a14;
    --panel: #15241c;
    --panel-2: #1c2f25;
    --accent: #4ade80;
    --accent-dim: #22c55e;
    --text: #e6f2ec;
    --muted: #94a89b;
    --neg: #f87171;
    --pos: #4ade80;
    --border: #25382c;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0;
    background: var(--bg);
    color: var(--text);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    -webkit-font-smoothing: antialiased;
  }
  .wrap { max-width: 1200px; margin: 0 auto; padding: 32px 24px 80px; }
  header { margin-bottom: 24px; }
  h1 { margin: 0 0 4px; font-size: 28px; font-weight: 700; letter-spacing: -0.02em; }
  .sub { color: var(--muted); font-size: 14px; }
  .grid { display: grid; gap: 16px; }
  .kpis { grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); margin-top: 16px; }
  .card {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 18px 20px;
  }
  .card h3 {
    margin: 0 0 8px;
    font-size: 12px;
    font-weight: 600;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.06em;
  }
  .card .val { font-size: 24px; font-weight: 700; letter-spacing: -0.01em; }
  .card .sub2 { font-size: 12px; color: var(--muted); margin-top: 4px; }
  .pos { color: var(--pos); }
  .neg { color: var(--neg); }
  .charts {
    grid-template-columns: 1.2fr 1fr;
    margin-top: 24px;
  }
  @media (max-width: 800px) { .charts { grid-template-columns: 1fr; } }
  .chart-card { background: var(--panel); border: 1px solid var(--border); border-radius: 12px; padding: 20px; }
  .chart-card h2 { margin: 0 0 16px; font-size: 16px; font-weight: 600; }
  .impact { grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); margin-top: 24px; }
  .impact .card .val { color: var(--accent); }
  .impact .card .sub2 { margin-top: 6px; }
  section { margin-top: 32px; }
  h2.section { font-size: 18px; font-weight: 600; margin: 0 0 12px; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  thead th {
    text-align: left;
    padding: 10px 12px;
    background: var(--panel-2);
    color: var(--muted);
    font-weight: 600;
    text-transform: uppercase;
    font-size: 11px;
    letter-spacing: 0.05em;
    border-bottom: 1px solid var(--border);
  }
  tbody td {
    padding: 10px 12px;
    border-bottom: 1px solid var(--border);
  }
  tbody tr:last-child td { border-bottom: none; }
  .num { text-align: right; font-variant-numeric: tabular-nums; }
  .table-wrap { background: var(--panel); border: 1px solid var(--border); border-radius: 12px; overflow: hidden; }
  .bar { display: inline-block; height: 6px; background: var(--accent-dim); border-radius: 3px; vertical-align: middle; margin-left: 6px; }
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1 id="title">Impact Portfolio</h1>
    <div class="sub" id="subtitle"></div>
  </header>

  <div class="grid kpis" id="kpis"></div>

  <div class="grid charts">
    <div class="chart-card">
      <h2>Sector allocation</h2>
      <canvas id="sectorChart"></canvas>
    </div>
    <div class="chart-card">
      <h2>Weighted ESG (0–100)</h2>
      <canvas id="esgChart"></canvas>
    </div>
  </div>

  <section>
    <h2 class="section">Annual impact</h2>
    <div class="grid impact" id="impact"></div>
  </section>

  <section>
    <h2 class="section">Holdings</h2>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Ticker</th><th>Name</th><th>Sector</th>
            <th class="num">Shares</th><th class="num">Cost</th>
            <th class="num">Value</th><th class="num">P/L</th>
            <th class="num">Return</th><th class="num">ESG</th>
          </tr>
        </thead>
        <tbody id="holdings"></tbody>
      </table>
    </div>
  </section>
</div>

<script>
const data = __DATA__;

function usd(n) { return "$" + n.toLocaleString("en-US", { minimumFractionDigits: 2, maximumFractionDigits: 2 }); }
function pct(n) { return (n >= 0 ? "+" : "") + n.toFixed(2) + "%"; }
function num(n, d) { return n.toLocaleString("en-US", { minimumFractionDigits: d, maximumFractionDigits: d }); }
function signClass(n) { return n >= 0 ? "pos" : "neg"; }

document.getElementById("title").textContent = data.name;
document.getElementById("subtitle").textContent = data.holdings.length + " holdings • impact-weighted view";

const kpis = [
  { label: "Cost basis",   val: usd(data.totals.cost),  sub: "" },
  { label: "Market value", val: usd(data.totals.value), sub: "" },
  { label: "Unrealized P/L", val: usd(data.totals.gain), sub: pct(data.totals.return_pct), cls: signClass(data.totals.gain) },
  { label: "Weighted ESG", val: data.esg.ESG.toFixed(1), sub: "E " + data.esg.E.toFixed(1) + " · S " + data.esg.S.toFixed(1) + " · G " + data.esg.G.toFixed(1) },
];
document.getElementById("kpis").innerHTML = kpis.map(k =>
  `<div class="card"><h3>${k.label}</h3><div class="val ${k.cls || ""}">${k.val}</div><div class="sub2 ${k.cls || ""}">${k.sub}</div></div>`
).join("");

const impactCards = [
  { label: "CO₂e avoided",    val: num(data.impact.co2e_avoided_t, 1) + " t",     sub: num(data.per1k.co2e_avoided_t, 3) + " t per $1k" },
  { label: "Clean energy",    val: num(data.impact.clean_energy_mwh, 1) + " MWh", sub: num(data.per1k.clean_energy_mwh, 3) + " MWh per $1k" },
  { label: "People served",   val: num(data.impact.people_served, 0),             sub: num(data.per1k.people_served, 2) + " per $1k" },
  { label: "Jobs supported",  val: num(data.impact.jobs_supported, 0),            sub: num(data.per1k.jobs_supported, 3) + " per $1k" },
];
document.getElementById("impact").innerHTML = impactCards.map(k =>
  `<div class="card"><h3>${k.label}</h3><div class="val">${k.val}</div><div class="sub2">${k.sub}</div></div>`
).join("");

document.getElementById("holdings").innerHTML = data.holdings.map(h => `
  <tr>
    <td><strong>${h.ticker}</strong></td>
    <td>${h.name}</td>
    <td>${h.sector}</td>
    <td class="num">${num(h.shares, 0)}</td>
    <td class="num">${usd(h.cost)}</td>
    <td class="num">${usd(h.value)}</td>
    <td class="num ${signClass(h.gain)}">${usd(h.gain)}</td>
    <td class="num ${signClass(h.return_pct)}">${pct(h.return_pct)}</td>
    <td class="num">${h.esg.toFixed(1)}</td>
  </tr>
`).join("");

const palette = ["#4ade80", "#22d3ee", "#a78bfa", "#fbbf24", "#fb7185", "#60a5fa", "#f472b6", "#34d399"];
const sectorEntries = Object.entries(data.sectors).sort((a, b) => b[1] - a[1]);

new Chart(document.getElementById("sectorChart"), {
  type: "doughnut",
  data: {
    labels: sectorEntries.map(([k]) => k),
    datasets: [{
      data: sectorEntries.map(([, v]) => v),
      backgroundColor: palette,
      borderColor: "#15241c",
      borderWidth: 2,
    }],
  },
  options: {
    plugins: {
      legend: { position: "right", labels: { color: "#e6f2ec", boxWidth: 12 } },
      tooltip: { callbacks: { label: c => c.label + ": " + c.parsed.toFixed(1) + "%" } },
    },
    cutout: "60%",
  },
});

new Chart(document.getElementById("esgChart"), {
  type: "bar",
  data: {
    labels: ["Environment", "Social", "Governance", "Combined"],
    datasets: [{
      data: [data.esg.E, data.esg.S, data.esg.G, data.esg.ESG],
      backgroundColor: ["#4ade80", "#22d3ee", "#a78bfa", "#fbbf24"],
      borderRadius: 6,
    }],
  },
  options: {
    indexAxis: "y",
    plugins: { legend: { display: false } },
    scales: {
      x: { min: 0, max: 100, ticks: { color: "#94a89b" }, grid: { color: "#25382c" } },
      y: { ticks: { color: "#e6f2ec" }, grid: { display: false } },
    },
  },
});
</script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
